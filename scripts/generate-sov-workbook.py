"""
Generate the Schedule of Values workbook used in Module 4, Lesson 1
(AI-Powered Formulas).

A hypothetical $32M marine construction project. The workbook contains:
  - Project header
  - Schedule of values: 21 line items, organized by CSI MasterFormat division
  - Pre-filled inputs (Total Contract Amount, Previously Billed, This Period)
  - Empty formula columns (Total Billed, % Complete, Remaining, Retention,
    Net Earned to Date, Status) — the trainee fills these in with AI's help
  - Division subtotal rows (also formula targets)
  - A grand total row

Run:
    python3 scripts/generate-sov-workbook.py
Output:
    public/downloads/sov-32m-project.xlsx
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter


# --------------------------------------------------------------------------- #
# Source data
# --------------------------------------------------------------------------- #

# (Item #, Division, Description, Total Contract Amount, Previously Billed, This Period)
LINE_ITEMS: list[tuple[int, str, str, int, int, int]] = [
    # Division 01 — General Requirements (3.0M)
    (1,  "01 — General Requirements", "Mobilization",                       1_200_000, 1_050_000,    150_000),
    (2,  "01 — General Requirements", "Bonds & Insurance",                    480_000,   480_000,          0),
    (3,  "01 — General Requirements", "Project Management (monthly)",         700_000,   180_000,     44_000),
    (4,  "01 — General Requirements", "Field Office / Temporary Facilities",  320_000,   270_000,     50_000),
    (5,  "01 — General Requirements", "Quality Control Program",              300_000,    75_000,     21_000),

    # Division 02 — Existing Conditions (2.4M)
    (6,  "02 — Existing Conditions",  "Demolition of existing structures",  1_800_000, 1_500_000,    210_000),
    (7,  "02 — Existing Conditions",  "Disposal & Hauling",                   600_000,   380_000,     70_000),

    # Division 03 — Concrete (4.0M)
    (8,  "03 — Concrete",             "Cast-in-place concrete (caps)",      2_400_000,   180_000,    180_000),
    (9,  "03 — Concrete",             "Reinforcing steel",                    880_000,   100_000,     93_600),
    (10, "03 — Concrete",             "Formwork",                             720_000,   100_000,    101_600),

    # Division 05 — Metals (5.2M)
    (11, "05 — Metals",               "Sheet pile fabrication",             4_200_000, 2_400_000,    330_000),
    (12, "05 — Metals",               "Sheet pile delivery to site",          480_000,   200_000,     88_000),
    (13, "05 — Metals",               "Miscellaneous structural steel",       520_000,    40_000,     22_400),

    # Division 31 — Earthwork (5.0M)
    (14, "31 — Earthwork",            "Dredging",                           3_600_000, 1_050_000,    318_000),
    (15, "31 — Earthwork",            "Backfill",                           1_400_000,         0,          0),

    # Division 33 — Utilities (0.4M)
    (16, "33 — Utilities",            "Stormwater management",                400_000,         0,          0),

    # Division 35 — Waterway and Marine Construction (12.0M)
    (17, "35 — Marine Construction",  "Sheet pile installation",            5_800_000,   400_000,    296_000),
    (18, "35 — Marine Construction",  "Pile driving",                       3_200_000,         0,    160_000),
    (19, "35 — Marine Construction",  "Fender system installation",         1_800_000,         0,          0),
    (20, "35 — Marine Construction",  "Mooring hardware",                     880_000,         0,          0),
    (21, "35 — Marine Construction",  "Demobilization & site restoration",    320_000,         0,          0),
]


# --------------------------------------------------------------------------- #
# Styling
# --------------------------------------------------------------------------- #

CASHMAN_GREEN = "2B6E2B"
CREAM = "FDFAF7"
BORDER_GRAY = "B7B7B7"
LIGHT_GRAY = "EFEFEF"
TINT_GREEN = "C8E9C8"
HEADLINE = "030303"
BODY = "272525"

THIN = Side(style="thin", color=BORDER_GRAY)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=18, bold=True, color=HEADLINE)
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color=BODY)
ROW_FONT = Font(name="Calibri", size=10, color=BODY)
SUBTOTAL_FONT = Font(name="Calibri", size=10, bold=True, color=HEADLINE)
TOTAL_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
GREEN_FILL = PatternFill("solid", fgColor=CASHMAN_GREEN)
TINT_FILL = PatternFill("solid", fgColor=TINT_GREEN)
SUBTOTAL_FILL = PatternFill("solid", fgColor=LIGHT_GRAY)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center")
CENTER = Alignment(horizontal="center", vertical="center")
CURRENCY = "$#,##0.00"


# --------------------------------------------------------------------------- #
# Workbook builder
# --------------------------------------------------------------------------- #


def build_workbook(output_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Schedule of Values"

    # Column widths
    widths = {
        "A": 5,    # Item #
        "B": 24,   # Division
        "C": 38,   # Description
        "D": 17,   # Total Contract Amount
        "E": 17,   # Previously Billed
        "F": 14,   # This Period
        "G": 16,   # Total Billed (formula)
        "H": 12,   # % Complete (formula)
        "I": 16,   # Remaining (formula)
        "J": 14,   # Retention 10% (formula)
        "K": 17,   # Net Earned to Date (formula)
        "L": 16,   # Status (formula)
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # ---------- Project header ----------
    ws["A1"] = "Cashman Marine — Hypothetical Project"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:L1")

    ws["A2"] = (
        "Schedule of Values  ·  Total Contract: $32,000,000  ·  "
        "Period billed to date: ~32%  ·  Retention rate: 10%"
    )
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:L2")

    ws["A3"] = (
        "Columns G, H, I, J, K, L need formulas. Use AI to generate them, then "
        "test with known values before trusting the totals."
    )
    ws["A3"].font = Font(name="Calibri", size=10, italic=True, color=CASHMAN_GREEN)
    ws.merge_cells("A3:L3")

    # Spacer
    ws.row_dimensions[4].height = 6

    # ---------- Column headers ----------
    headers = [
        "#",
        "Division",
        "Line Item Description",
        "Total Contract Amount",
        "Previously Billed",
        "This Period",
        "Total Billed (formula)",
        "% Complete (formula)",
        "Remaining (formula)",
        "Retention 10% (formula)",
        "Net Earned to Date (formula)",
        "Status (formula)",
    ]
    HEADER_ROW = 5
    for idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=HEADER_ROW, column=idx, value=h)
        cell.fill = GREEN_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )
        cell.border = BOX
    ws.row_dimensions[HEADER_ROW].height = 32

    # ---------- Line items ----------
    DATA_START = HEADER_ROW + 1
    for offset, (item, div, desc, total, prev, this) in enumerate(LINE_ITEMS):
        r = DATA_START + offset
        ws.cell(row=r, column=1, value=item).alignment = CENTER
        ws.cell(row=r, column=2, value=div).alignment = LEFT
        ws.cell(row=r, column=3, value=desc).alignment = LEFT
        ws.cell(row=r, column=4, value=total)
        ws.cell(row=r, column=5, value=prev)
        ws.cell(row=r, column=6, value=this)

        for c in (4, 5, 6):
            ws.cell(row=r, column=c).number_format = CURRENCY
            ws.cell(row=r, column=c).alignment = RIGHT
        # Empty formula cells (G–L) get the currency format up front so when
        # the trainee fills in formulas, the result displays correctly.
        for c in (7, 9, 10, 11):
            ws.cell(row=r, column=c).number_format = CURRENCY
        ws.cell(row=r, column=8).number_format = "0.00%"
        ws.cell(row=r, column=12).alignment = CENTER

        for c in range(1, 13):
            ws.cell(row=r, column=c).font = ROW_FONT
            ws.cell(row=r, column=c).border = BOX

    DATA_END = DATA_START + len(LINE_ITEMS) - 1

    # ---------- Spacer ----------
    SPACER = DATA_END + 1
    ws.row_dimensions[SPACER].height = 8

    # ---------- Division subtotals ----------
    # Trainee writes SUMIFS formulas here.
    SUB_START = SPACER + 1
    ws.cell(row=SUB_START, column=2, value="DIVISION SUBTOTALS").font = SUBTOTAL_FONT
    ws.cell(row=SUB_START, column=2).fill = TINT_FILL
    ws.merge_cells(start_row=SUB_START, start_column=2, end_row=SUB_START, end_column=12)
    sub_header = ws.cell(row=SUB_START, column=2)
    sub_header.alignment = Alignment(horizontal="left", vertical="center")

    divisions = [
        "01 — General Requirements",
        "02 — Existing Conditions",
        "03 — Concrete",
        "05 — Metals",
        "31 — Earthwork",
        "33 — Utilities",
        "35 — Marine Construction",
    ]
    SUB_DATA_START = SUB_START + 1
    for offset, div in enumerate(divisions):
        r = SUB_DATA_START + offset
        ws.cell(row=r, column=2, value=div).alignment = LEFT
        ws.cell(row=r, column=3, value="(SUMIFS by division — formula needed)").alignment = LEFT
        ws.cell(row=r, column=3).font = Font(
            name="Calibri", size=10, italic=True, color="888888"
        )
        for c in (4, 5, 6, 7, 9, 10, 11):
            ws.cell(row=r, column=c).number_format = CURRENCY
        ws.cell(row=r, column=8).number_format = "0.00%"
        for c in range(1, 13):
            ws.cell(row=r, column=c).font = SUBTOTAL_FONT
            ws.cell(row=r, column=c).fill = SUBTOTAL_FILL
            ws.cell(row=r, column=c).border = BOX
    SUB_DATA_END = SUB_DATA_START + len(divisions) - 1

    # ---------- Spacer ----------
    SPACER2 = SUB_DATA_END + 1
    ws.row_dimensions[SPACER2].height = 8

    # ---------- Grand total ----------
    GRAND = SPACER2 + 1
    ws.cell(row=GRAND, column=2, value="PROJECT TOTAL").alignment = LEFT
    ws.cell(row=GRAND, column=3, value="(SUM across all line items — formula needed)").alignment = LEFT
    ws.cell(row=GRAND, column=3).font = Font(
        name="Calibri", size=10, italic=True, color="EEEEEE"
    )
    for c in (4, 5, 6, 7, 9, 10, 11):
        ws.cell(row=GRAND, column=c).number_format = CURRENCY
    ws.cell(row=GRAND, column=8).number_format = "0.00%"
    for c in range(1, 13):
        cell = ws.cell(row=GRAND, column=c)
        cell.fill = GREEN_FILL
        cell.font = TOTAL_FONT
        cell.border = BOX

    # ---------- Notes ----------
    NOTES_ROW = GRAND + 3
    ws.cell(row=NOTES_ROW, column=2, value="Notes for the formula columns:").font = Font(
        name="Calibri", size=11, bold=True, color=HEADLINE
    )
    ws.merge_cells(
        start_row=NOTES_ROW, start_column=2, end_row=NOTES_ROW, end_column=12
    )

    notes = [
        "Total Billed (G)         =  Previously Billed (E) + This Period (F)",
        "% Complete (H)           =  Total Billed (G) / Total Contract Amount (D)",
        "Remaining (I)            =  Total Contract Amount (D) − Total Billed (G)",
        "Retention 10% (J)        =  Total Billed (G) × 10%",
        "Net Earned to Date (K)   =  Total Billed (G) − Retention (J)",
        "Status (L)               =  COMPLETE | NEAR COMPLETE (≥90%) | IN PROGRESS (>0%) | NOT STARTED (0%)",
        "Division subtotals       =  SUMIFS by division (one row per division)",
        "Project Total            =  SUM of each numeric column across all line items",
    ]
    for offset, note in enumerate(notes):
        r = NOTES_ROW + 1 + offset
        ws.cell(row=r, column=2, value=note).font = Font(
            name="Consolas", size=10, color=BODY
        )
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
        ws.row_dimensions[r].height = 16

    # Freeze the header so the trainee can scroll the data without losing context
    ws.freeze_panes = ws.cell(row=DATA_START, column=4)

    # Save
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


def main() -> None:
    repo_root = Path(__file__).resolve().parent.parent
    output = repo_root / "public" / "downloads" / "sov-32m-project.xlsx"
    build_workbook(output)
    print(f"Generated {output.relative_to(repo_root)}")
    # Verify totals
    total_contract = sum(item[3] for item in LINE_ITEMS)
    total_billed = sum(item[4] + item[5] for item in LINE_ITEMS)
    pct = total_billed / total_contract * 100
    print(f"  Lines: {len(LINE_ITEMS)}")
    print(f"  Total contract: ${total_contract:,}")
    print(f"  Total billed:   ${total_billed:,}  ({pct:.2f}%)")


if __name__ == "__main__":
    main()
