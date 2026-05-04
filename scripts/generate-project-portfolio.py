"""
Generate the Cashman project portfolio used in Module 4, Lesson 2
(Data Analysis with AI).

A 32-row dataset covering four hypothetical work types Cashman / its
subsidiaries actually do:
    - Dredging (Cashman marine)
    - Pile Driving (Cashman marine)
    - IPC Lydon (industrial power / mechanical contracting)
    - Preload Cryogenics (LNG / cryogenic tank construction)

Per project: name, type, region, contract value, final cost, planned
duration (months), actual duration (months), # change orders, change order
$, client satisfaction (1-5), PM, status. Distribution is intentionally
realistic — some projects under budget, some way over, with a pattern the
LLM can find when asked "which type goes over budget and why."

Run:
    python3 scripts/generate-project-portfolio.py
Output:
    public/downloads/cashman-project-portfolio.xlsx
"""

from __future__ import annotations

import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side


# --------------------------------------------------------------------------- #
# Source data
# --------------------------------------------------------------------------- #
# Tuple shape: (Project Name, Type, Region, Contract $, Final Cost $,
#               Planned Months, Actual Months, # Change Orders, CO Total $,
#               Satisfaction 1-5, PM, Status)
#
# Patterns intentionally baked in:
#   - Dredging projects: usually on budget, occasional over from weather
#   - Pile Driving: tight margin, mostly on time
#   - IPC Lydon: variable; complex industrial scope tends to grow
#   - Preload Cryogenics: heavily over budget when scope expands; a few
#     clean wins. Long durations.
PROJECTS: list[tuple] = [
    # --- Dredging (8) ---
    ("Boston Inner Harbor Channel Maint",       "Dredging",            "Boston, MA",        4_200_000,  4_120_000,  4,   4,  1,   80_000,  5, "Bobby M.",  "Complete"),
    ("New Bedford Harbor Deepening",            "Dredging",            "New Bedford, MA",  12_400_000, 13_900_000, 11,  13,  4,  920_000,  4, "Bobby M.",  "Complete"),
    ("Providence River Channel Restoration",    "Dredging",            "Providence, RI",    6_800_000,  6_650_000,  6,   6,  2,  140_000,  5, "Sarah K.",  "Complete"),
    ("Portland Berth 7 Maintenance",            "Dredging",            "Portland, ME",      3_100_000,  3_090_000,  3,   3,  0,        0,  5, "Bobby M.",  "Complete"),
    ("Long Island Sound Disposal Site",         "Dredging",            "Long Island, NY",   9_600_000, 11_200_000,  8,  11,  6, 1_350_000, 3, "Sarah K.",  "Complete"),
    ("Chesapeake Approach Channel",             "Dredging",            "Norfolk, VA",      18_500_000, 18_300_000, 14,  14,  3,  280_000,  5, "Bobby M.",  "Complete"),
    ("Conanicut Island Marina",                 "Dredging",            "Jamestown, RI",     2_400_000,  2_380_000,  3,   3,  1,   55_000,  4, "Sarah K.",  "Complete"),
    ("Sandwich Town Marine Maint",              "Dredging",            "Sandwich, MA",      3_900_000,  4_050_000,  4,   5,  2,  165_000,  4, "Bobby M.",  "Complete"),

    # --- Pile Driving (8) ---
    ("Quincy Shipyard Wharf Recap",             "Pile Driving",        "Quincy, MA",        7_800_000,  7_900_000,  7,   8,  2,  145_000,  4, "Mike R.",   "Complete"),
    ("Galveston Wharf Rehabilitation",          "Pile Driving",        "Galveston, TX",     4_200_000,  4_320_000,  5,   6,  3,  180_000,  4, "Mike R.",   "Complete"),
    ("Cape Cod Canal Fender System",            "Pile Driving",        "Bourne, MA",        2_900_000,  2_870_000,  4,   4,  1,   45_000,  5, "Wes T.",    "Complete"),
    ("Hyannis Ferry Terminal Recap",            "Pile Driving",        "Hyannis, MA",       5_100_000,  5_080_000,  6,   6,  1,   60_000,  5, "Mike R.",   "Complete"),
    ("Throgs Neck Pier Reconstruction",         "Pile Driving",        "Bronx, NY",         9_800_000, 10_400_000,  9,  11,  4,  680_000,  4, "Wes T.",    "Complete"),
    ("Salem Wharf Strengthening",               "Pile Driving",        "Salem, MA",         3_300_000,  3_280_000,  4,   4,  0,        0,  5, "Mike R.",   "Complete"),
    ("New London Sub Base Pier",                "Pile Driving",        "New London, CT",   14_500_000, 14_900_000, 12,  13,  3,  520_000,  4, "Wes T.",    "Complete"),
    ("Cohasset Town Pier Replacement",          "Pile Driving",        "Cohasset, MA",      1_900_000,  1_950_000,  3,   3,  1,   42_000,  5, "Mike R.",   "Complete"),

    # --- IPC Lydon (industrial process / mechanical) (8) ---
    ("Pawtucket WWTP Mechanical Upgrade",       "IPC Lydon",           "Pawtucket, RI",     8_200_000,  9_100_000, 10,  13,  6,  860_000,  3, "Peter S.",  "Complete"),
    ("Holyoke Cogen Boiler Retrofit",           "IPC Lydon",           "Holyoke, MA",       6_400_000,  7_900_000,  8,  12,  9, 1_350_000, 2, "Peter S.",  "Complete"),
    ("Chelsea Pump Station Mechanical",         "IPC Lydon",           "Chelsea, MA",       4_800_000,  5_100_000,  6,   7,  4,  280_000,  4, "Karen H.",  "Complete"),
    ("Stoughton Process Piping Install",        "IPC Lydon",           "Stoughton, MA",     2_900_000,  2_960_000,  4,   5,  2,   75_000,  4, "Karen H.",  "Complete"),
    ("Salem Steam Plant Repipe",                "IPC Lydon",           "Salem, MA",         5_400_000,  6_200_000,  7,   9,  5,  720_000,  3, "Peter S.",  "Complete"),
    ("Gardner Power Plant Boiler Tube",         "IPC Lydon",           "Gardner, MA",       3_700_000,  3_650_000,  5,   5,  1,   55_000,  5, "Karen H.",  "Complete"),
    ("Worcester Hospital Mech Phase 2",         "IPC Lydon",           "Worcester, MA",    11_800_000, 13_400_000, 12,  16,  8, 1_450_000, 3, "Peter S.",  "Complete"),
    ("Topeka Industrial Steam (Kansas)",        "IPC Lydon",           "Topeka, KS",       18_500_000, 22_900_000, 14,  19, 12, 3_900_000, 2, "Peter S.",  "Complete"),

    # --- Preload Cryogenics (LNG / cryogenic tank construction) (8) ---
    ("Tampa LNG Storage Tank Phase 1",          "Preload Cryogenics",  "Tampa, FL",        42_000_000, 41_500_000, 22,  22,  2,  380_000,  5, "Wes T.",    "Complete"),
    ("Quintana Island LNG Tank Foundation",     "Preload Cryogenics",  "Quintana, TX",     38_500_000, 47_200_000, 20,  27, 11, 7_400_000, 2, "Wes T.",    "Complete"),
    ("Sabine Pass LNG Tank Recap",              "Preload Cryogenics",  "Sabine Pass, LA",  61_000_000, 75_500_000, 28,  34, 14, 12_900_000,2, "Karen H.",  "Complete"),
    ("Cove Point LNG Tank Modifications",       "Preload Cryogenics",  "Lusby, MD",        29_400_000, 31_200_000, 18,  20,  6, 1_500_000, 4, "Wes T.",    "Complete"),
    ("Freeport LNG Storage Tank #4",            "Preload Cryogenics",  "Freeport, TX",     54_000_000, 53_800_000, 24,  24,  3,  600_000,  5, "Karen H.",  "Complete"),
    ("Elba Island LNG Tank Refurb",             "Preload Cryogenics",  "Elba Island, GA",  19_200_000, 23_500_000, 14,  18,  9, 3_800_000, 2, "Wes T.",    "Complete"),
    ("Plaquemines Parish LNG Tank Found",       "Preload Cryogenics",  "Plaquemines, LA",  45_000_000, 51_400_000, 22,  27, 10, 5_500_000, 3, "Karen H.",  "Complete"),
    ("Calcasieu Pass LNG Phase 3",              "Preload Cryogenics",  "Cameron, LA",      72_000_000, 89_500_000, 30,  38, 16, 15_200_000,2, "Wes T.",    "In Progress"),
]


HEADERS = [
    "Project Name",
    "Project Type",
    "Region",
    "Contract Value",
    "Final Cost",
    "Cost Variance",
    "Planned Duration (months)",
    "Actual Duration (months)",
    "Schedule Variance (months)",
    "# Change Orders",
    "Change Order Total",
    "Client Satisfaction (1-5)",
    "Project Manager",
    "Status",
]


# --------------------------------------------------------------------------- #
# Styling
# --------------------------------------------------------------------------- #

CASHMAN_GREEN = "2B6E2B"
LIGHT_GRAY = "EFEFEF"
BORDER_GRAY = "B7B7B7"
HEADLINE = "030303"
BODY = "272525"

THIN = Side(style="thin", color=BORDER_GRAY)
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEAD_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="Calibri", size=18, bold=True, color=HEADLINE)
SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color=BODY)
ROW_FONT = Font(name="Calibri", size=10, color=BODY)
GREEN_FILL = PatternFill("solid", fgColor=CASHMAN_GREEN)
ALT_FILL = PatternFill("solid", fgColor="F7F7F2")
CURRENCY = "$#,##0"
RIGHT = Alignment(horizontal="right", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
CENTER = Alignment(horizontal="center", vertical="center")


def build(output_path: Path) -> None:
    random.seed(42)  # stable output for diffs / regen
    wb = Workbook()
    ws = wb.active
    ws.title = "Project Portfolio"

    widths = {
        "A": 38, "B": 18, "C": 18, "D": 16, "E": 16, "F": 14,
        "G": 12, "H": 12, "I": 12, "J": 10, "K": 16, "L": 14,
        "M": 14, "N": 14,
    }
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    # Header banner
    ws["A1"] = "Cashman Companies — Project Portfolio (last 5 years)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:N1")

    ws["A2"] = (
        "32 completed (or in-progress) projects across Dredging, Pile Driving, "
        "IPC Lydon, and Preload Cryogenics. Use this dataset to ask AI which "
        "project type tends to go over budget and why."
    )
    ws["A2"].font = SUBTITLE_FONT
    ws.merge_cells("A2:N2")
    ws.row_dimensions[2].height = 28
    ws["A2"].alignment = LEFT

    # spacer
    ws.row_dimensions[3].height = 6

    # Column headers
    HEADER_ROW = 4
    for idx, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=HEADER_ROW, column=idx, value=h)
        cell.fill = GREEN_FILL
        cell.font = HEAD_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = BOX
    ws.row_dimensions[HEADER_ROW].height = 36

    # Data rows
    DATA_START = HEADER_ROW + 1
    for offset, p in enumerate(PROJECTS):
        r = DATA_START + offset
        (name, type_, region, contract, final, planned, actual,
         co_count, co_total, sat, pm, status) = p
        cost_variance = final - contract
        sched_variance = actual - planned

        values = [
            name, type_, region, contract, final, cost_variance,
            planned, actual, sched_variance, co_count, co_total, sat, pm, status,
        ]
        for c, v in enumerate(values, start=1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.font = ROW_FONT
            cell.border = BOX
            if r % 2 == 0:
                cell.fill = ALT_FILL
            if c in (4, 5, 6, 11):  # currency columns
                cell.number_format = CURRENCY
                cell.alignment = RIGHT
            elif c in (7, 8, 9, 10, 12):
                cell.alignment = CENTER
            else:
                cell.alignment = LEFT

    DATA_END = DATA_START + len(PROJECTS) - 1

    # Freeze header row + project name column
    ws.freeze_panes = ws.cell(row=DATA_START, column=2)

    # Notes / legend
    NOTES_ROW = DATA_END + 3
    ws.cell(row=NOTES_ROW, column=1, value="Notes").font = Font(
        name="Calibri", size=11, bold=True, color=HEADLINE
    )
    ws.merge_cells(start_row=NOTES_ROW, start_column=1, end_row=NOTES_ROW, end_column=14)

    notes = [
        "• Cost Variance = Final Cost − Contract Value (positive = over budget)",
        "• Schedule Variance = Actual Duration − Planned Duration (positive = late)",
        "• Client Satisfaction is on a 1 (worst) to 5 (best) scale, captured at project closeout",
        "• Status \"In Progress\" rows have a Final Cost that's the latest forecast, not actual",
        "• All figures are hypothetical and for training only",
    ]
    for offset, note in enumerate(notes):
        r = NOTES_ROW + 1 + offset
        ws.cell(row=r, column=1, value=note).font = Font(
            name="Calibri", size=10, color=BODY
        )
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=14)
        ws.row_dimensions[r].height = 16

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))


def main() -> None:
    repo_root = Path(__file__).resolve().parent.parent
    output = repo_root / "public" / "downloads" / "cashman-project-portfolio.xlsx"
    build(output)
    print(f"Generated {output.relative_to(repo_root)}")
    # Verify dataset stats so the summary in the lesson stays accurate.
    by_type: dict[str, list[tuple]] = {}
    for p in PROJECTS:
        by_type.setdefault(p[1], []).append(p)
    print(f"  Projects: {len(PROJECTS)}")
    for t, rows in sorted(by_type.items()):
        contract = sum(r[3] for r in rows)
        final = sum(r[4] for r in rows)
        co = sum(r[8] for r in rows)
        avg_sat = sum(r[9] for r in rows) / len(rows)
        over_pct = (final - contract) / contract * 100
        print(
            f"  {t:24s} n={len(rows)}  contract=${contract:>13,}  "
            f"final=${final:>13,}  delta={over_pct:+.1f}%  COs={co:>2}  avg_sat={avg_sat:.1f}"
        )


if __name__ == "__main__":
    main()
